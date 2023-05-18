import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Font, Alignment
import re
from deepdiff import DeepDiff
from datetime import date
# Variable Input
workBookPath = r'../ComparisonCustLib.xlsx'
firstWsName = "Analyzed"
secondWsName = "FuzzyTestApr"
today = date.today().strftime("%d%m%Y")

# Fill Input
wb = openpyxl.load_workbook(workBookPath)
ws1 = wb[firstWsName]
ws2 = wb[secondWsName]
wsDataSet1 = {
    "maxRows": len(ws1['B']),
    "columnA": 2, #FileName
    "columnB": 7, #RuleName
    "columnC": 6, #Priority
    "columnD": 13, #Analysis
    "columnE": 8,  #Variable
}
wsDataSet2 = {
    "maxRows": len(ws2['B']),
    "columnA": 2, #FileName
    "columnB": 6, #RuleName
    "columnC": 5, #Priority
    "columnD": 11, #Analysis
    "columnE": 7,  #Variable
}
dic1 = {}
dic2 = {}
copyDic1 = {}
copyDic2 = {}
analysisDic1 = {}
unresolvedRetrivedDic1 = {}

def CreateDataSet(ws, wsDataSet):
    dic = {}
    copyDic = {}
    flag = False
    firstRow = 0
    lastRow = 0
    for nRow in range(2, wsDataSet["maxRows"] + 1):
        cellValPrev = ws.cell(row=nRow - 1, column=wsDataSet["columnA"])
        cellValCurr = ws.cell(row=nRow, column=wsDataSet["columnA"])
        priorityRule = ws.cell(row=nRow - 1, column=wsDataSet["columnC"])
        if priorityRule.value == "high":
            if cellValPrev.value == cellValCurr.value and not flag:
                flag = True
                firstRow = nRow - 1
            elif cellValPrev.value != cellValCurr.value:
                if flag:
                    lastRow = nRow - 1
                    flag = False
                    if cellValPrev.value in dic:
                        dic[cellValPrev.value].append([firstRow, lastRow])
                    else:
                        listAdd = [[firstRow, lastRow]]
                        dic[cellValPrev.value] = listAdd
                else:
                    if cellValPrev.value in dic:
                        dic[cellValPrev.value].append([nRow - 1, nRow - 1])
                    else:
                        listAdd = [[nRow - 1, nRow - 1]]
                        dic[cellValPrev.value] = listAdd
    # Store current dictionary for further usage
    for key, value in dic.items():
        copyDic[key] = value
    # Add rules into the related files
    subDic = {}
    for key, value in dic.items():
        for elem in value:
            if elem[0] != elem[1]:
                for nRow2 in range(elem[0], elem[1] + 1):
                    cellVarB = ws.cell(row=nRow2, column=wsDataSet["columnB"])
                    if cellVarB.value not in subDic:
                        subDic.setdefault(cellVarB.value, 1)
                    else:
                        subDic[cellVarB.value] += 1
            else:
                nRow2 = elem[0]
                cellVarB = ws.cell(row=nRow2, column=wsDataSet["columnB"])
                subDic.setdefault(cellVarB.value, 1)
        dic[key] = subDic
        subDic = {}
    return dic, copyDic

def FillDataAnalysis(retrivedDic, filledDic , wsRetr, wsFilled, wsDataSetRetr, wsDataSetFilled):
    analysisDic = {}
    for key, value in retrivedDic.items():
        for elem in value:
            if elem[0] != elem[1]:
                for nRow2 in range(elem[0], elem[1] + 1):
                    cellVarD = wsRetr.cell(row=nRow2, column=wsDataSetRetr["columnD"]) #Analysis
                    cellVarB = wsRetr.cell(row=nRow2, column=wsDataSetRetr["columnB"]) #Rules
                    cellVarE = wsRetr.cell(row=nRow2, column=wsDataSetRetr["columnE"]) #Variables
                    if key in analysisDic:
                        analysisDic[key].append({cellVarB.value: {cellVarE.value: cellVarD.value}})
                    else:
                        analysisDic[key] = [{cellVarB.value: {cellVarE.value: cellVarD.value}}] # {Rule : {Var: Analysis}}
            else:
                nRow2 = elem[0]
                cellVarD = wsRetr.cell(row=nRow2, column=wsDataSetRetr["columnD"]) #Analysis
                cellVarB = wsRetr.cell(row=nRow2, column=wsDataSetRetr["columnB"])  # Rules
                cellVarE = wsRetr.cell(row=nRow2, column=wsDataSetRetr["columnE"])  # Variables
                if key in analysisDic:
                    analysisDic[key].append({cellVarB.value: {cellVarE.value: cellVarD.value}})
                else:
                    analysisDic[key] = [{cellVarB.value: {cellVarE.value: cellVarD.value}}]
    # Store current dictionary for further usage
    TestDic = analysisDic.copy()
    CreateText(TestDic, "TestDic1")
    # Fill in needed wb
    for key, value in filledDic.items():
        for elem in value:
            if elem[0] != elem[1]:
                for nRow2 in range(elem[0], elem[1] + 1):
                    cellVarD = wsFilled.cell(row=nRow2, column=wsDataSetFilled["columnD"]) #Analysis
                    cellVarB = wsFilled.cell(row=nRow2, column=wsDataSetFilled["columnB"]) # Rules
                    cellVarE = wsFilled.cell(row=nRow2, column=wsDataSetFilled["columnE"]) # Variables
                    if key in analysisDic:
                        # Need to check whether the analysis queue is empty or not after popping its elements
                        if len(analysisDic[key]):
                            if cellVarB.value in analysisDic[key][0]:
                                if cellVarE.value in analysisDic[key][0][cellVarB.value]:
                                    cellVarD.value = analysisDic[key][0][cellVarB.value][cellVarE.value]
                                    analysisDic[key].pop(0)
                                else:
                                    for remainIdx in range(1, len(analysisDic[key])):
                                        if cellVarB.value in analysisDic[key][remainIdx]:
                                            if cellVarE.value in analysisDic[key][remainIdx][cellVarB.value]:
                                                cellVarD.value = analysisDic[key][remainIdx][cellVarB.value][cellVarE.value]
                                                analysisDic[key].pop(remainIdx)
                                                break
                                    if cellVarD.value == "NOK":
                                        cellVarD.value = "Need to fill an analysis 1"
                                        analysisDic[key].append(analysisDic[key].pop(0))
                            else:
                                for idx in range(len(analysisDic[key])):
                                    if idx < len(analysisDic[key]):
                                        if cellVarB.value in analysisDic[key][idx]:
                                            # Need to find the first matching rule in the analysisDic then do the same steps as from line 165 to line 188
                                            # cellVarD.value = "Need to fill an analysis 2"
                                            if cellVarE.value in analysisDic[key][idx][cellVarB.value]:
                                                cellVarD.value = analysisDic[key][idx][cellVarB.value][cellVarE.value]
                                            else:
                                                # should halt this outer iteration until the current unresolved cell is filled
                                                for remainIdx in range(len(analysisDic[key])):
                                                    if cellVarB.value in analysisDic[key][remainIdx]:
                                                        if cellVarE.value in analysisDic[key][remainIdx][cellVarB.value]:
                                                            cellVarD.value = analysisDic[key][remainIdx][cellVarB.value][
                                                                cellVarE.value]
                                                            analysisDic[key].pop(remainIdx)
                                                            break
                                                if cellVarD.value == "NOK":
                                                    cellVarD.value = "Need to fill an analysis 1"
                                                    analysisDic[key].append(analysisDic[key].pop(idx))
                        else:
                            cellVarD.value = "Need to fill an analysis 2"
                        # else the queue is empty
                    else:
                        cellVarD.value = "Need to fill an analysis 3"
            else:
                nRow2 = elem[0]
                cellVarD = wsFilled.cell(row=nRow2, column=wsDataSetFilled["columnD"])  # Analysis
                cellVarB = wsFilled.cell(row=nRow2, column=wsDataSetFilled["columnB"])  # Rules
                cellVarE = wsFilled.cell(row=nRow2, column=wsDataSetFilled["columnE"])  # Variables
                if key in analysisDic:
                    # Need to check whether the analysis queue is empty or not after popping its elements
                    if len(analysisDic[key]):
                        if cellVarB.value in analysisDic[key][0]:
                            if cellVarE.value in analysisDic[key][0][cellVarB.value]:
                                cellVarD.value = analysisDic[key][0][cellVarB.value][cellVarE.value]
                            else:
                                # should halt this outer iteration until the current unresolved cell is filled
                                for remainIdx in range(1, len(analysisDic[key])):
                                    if cellVarB.value in analysisDic[key][remainIdx]:
                                        if cellVarE.value in analysisDic[key][remainIdx][cellVarB.value]:
                                            cellVarD.value = analysisDic[key][remainIdx][cellVarB.value][cellVarE.value]
                                            analysisDic[key].pop(remainIdx)
                                            break
                                if cellVarD.value == "NOK":
                                    cellVarD.value = "Need to fill an analysis 1"
                                    analysisDic[key].append(analysisDic[key].pop(0))
                        else:
                            for idx in range(len(analysisDic[key])):
                                if idx < len(analysisDic[key]):
                                    if cellVarB.value in analysisDic[key][idx]:
                                    # Need to find the first matching rule in the analysisDic then do the same steps as from line 165 to line 188
                                    #cellVarD.value = "Need to fill an analysis 2"
                                        if cellVarE.value in analysisDic[key][idx][cellVarB.value]:
                                            cellVarD.value = analysisDic[key][idx][cellVarB.value][cellVarE.value]
                                        else:
                                            # should halt this outer iteration until the current unresolved cell is filled
                                            for remainIdx in range(len(analysisDic[key])):
                                                if cellVarB.value in analysisDic[key][remainIdx]:
                                                    if cellVarE.value in analysisDic[key][remainIdx][cellVarB.value]:
                                                        cellVarD.value = analysisDic[key][remainIdx][cellVarB.value][
                                                            cellVarE.value]
                                                        analysisDic[key].pop(remainIdx)
                                                        break
                                            if cellVarD.value == "NOK":
                                                cellVarD.value = "Need to fill an analysis 1"
                                                analysisDic[key].append(analysisDic[key].pop(idx))
                    else:
                        cellVarD.value = "Need to fill an analysis 2"
                    # else the queue is empty
                else:
                    cellVarD.value = "Need to fill an analysis 3"
    return analysisDic


def CreateText(dic, filename):
    with open(f'../{filename}.txt', 'w') as f:
        for x, y in dic.items():
            text = x + '\t' + ':' + str(y)
            f.write(text)
            f.write('\n')

def BorderCell(cell):
    pink = "00FF00FF"
    green = "00008000"
    black = "000e2e2a"
    thick = Side(border_style="thick", color=black)
    double = Side(border_style="double", color=green)
    cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)

def FillBackGroundColor(cell):
    yellow = "00FFFF00"
    cell.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type="solid")

def FontCell(cell):
    cell.font = Font(name="Arial", size=14, color="00FF0000")

def InsertData(data, ws):
    nCol = 1
    for key, value in data.items():
        if key == "dictionary_item_added":
            key = firstWsName + " removed"
        elif key == "dictionary_item_removed":
            key = firstWsName + " added"
        else:
            key = f"Values_changed : (new_value: {secondWsName}; old_value: {firstWsName})"
        cell = ws.cell(row=1, column=nCol, value=key)
        # Formatting cell
        BorderCell(cell)
        FillBackGroundColor(cell)
        FontCell(cell)
        value = str(value)
        value = value.replace('root', '')
        if not re.findall("value.*", key):
            value = value.split(',')
        else:
            value = value.split("},")
        nRow = 2
        for eachVal in value:
            cell = ws.cell(row=nRow, column=nCol, value=eachVal)
            BorderCell(cell)
            nRow += 1
        nCol += 1

def generateDiffWorksheet(wsName):
    try:
        ws3 = wb.create_sheet(f"{wsName}")
        InsertData(diff, ws3)
        wb.save(f"../Result_{today}.xlsx")
        print("Saving the current workbook successfully")
    except PermissionError:
        print("Please close workbook before saving")

if __name__ == "__main__":
    i = 0
    dic1, copyDic1 = CreateDataSet(ws1, wsDataSet1)
    dic2, copyDic2 = CreateDataSet(ws2, wsDataSet2)
    analysisDic1 = FillDataAnalysis(copyDic1, copyDic2, ws1, ws2, wsDataSet1, wsDataSet2)
    diff = DeepDiff(dic1, dic2)

    CreateText(copyDic1, firstWsName)
    CreateText(copyDic2, secondWsName)
    CreateText(analysisDic1, "analysisDic1")

    #generateDiffWorksheet(f"Diff_{firstWsName}_{secondWsName}")


